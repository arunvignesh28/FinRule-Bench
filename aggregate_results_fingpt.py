#!/usr/bin/env python3
"""
aggregate_results_fingpt.py — Collect all FinGPT eval metrics and produce a summary table.

Reads *_fingpt_eval_metrics.txt files from all task subdirectories and compiles
a consolidated Excel/CSV summary across difficulty × statement type × prompt strategy.

Run from the repo root:
    python aggregate_results_fingpt.py
"""
import re
import csv
from pathlib import Path
from collections import defaultdict

REPO_ROOT = Path(__file__).resolve().parent

# ── Metric file patterns (relative to each script's directory) ────────────────
EASY_METRIC_FILES = {
    ("easy", "BS", "zero_shot"):    "task/easy/BS/FinGPT/zero_shot_easy_bs_fingpt_eval_metrics.txt",
    ("easy", "BS", "few_shot"):     "task/easy/BS/FinGPT/fewshot_easy_bs_fingpt_eval_metrics.txt",
    ("easy", "BS", "counterfactual"): "task/easy/BS/FinGPT/fewshot_cf_easy_bs_fingpt_eval_metrics.txt",
    ("easy", "CF", "zero_shot"):    "task/easy/CF/FinGPT/zero_shot_easy_cf_fingpt_eval_metrics.txt",
    ("easy", "CF", "few_shot"):     "task/easy/CF/FinGPT/fewshot_easy_cf_fingpt_eval_metrics.txt",
    ("easy", "CF", "counterfactual"): "task/easy/CF/FinGPT/fewshot_cf_easy_cf_fingpt_eval_metrics.txt",
    ("easy", "SE", "zero_shot"):    "task/easy/SE/FinGPT/zero_shot_easy_se_fingpt_eval_metrics.txt",
    ("easy", "SE", "few_shot"):     "task/easy/SE/FinGPT/fewshot_easy_se_fingpt_eval_metrics.txt",
    ("easy", "SE", "counterfactual"): "task/easy/SE/FinGPT/fewshot_cf_easy_se_fingpt_eval_metrics.txt",
    ("easy", "SI", "zero_shot"):    "task/easy/SI/FinGPT/zero_shot_easy_si_fingpt_eval_metrics.txt",
    ("easy", "SI", "few_shot"):     "task/easy/SI/FinGPT/fewshot_easy_si_fingpt_eval_metrics.txt",
    ("easy", "SI", "counterfactual"): "task/easy/SI/FinGPT/fewshot_cf_easy_si_fingpt_eval_metrics.txt",
}

MEDIUM_METRIC_FILES = {
    ("medium", "BS", "zero_shot"):    "task/medium/BS/FinGPT/zero_shot_medium_bs_fingpt_eval_metrics.txt",
    ("medium", "BS", "few_shot"):     "task/medium/BS/FinGPT/few_shot_medium_bs_fingpt_eval_metrics.txt",
    ("medium", "BS", "counterfactual"): "task/medium/BS/FinGPT/few_shot_cf_medium_bs_fingpt_eval_metrics.txt",
    ("medium", "CF", "zero_shot"):    "task/medium/CF/FinGPT/zero_shot_medium_cf_fingpt_eval_metrics.txt",
    ("medium", "CF", "few_shot"):     "task/medium/CF/FinGPT/few_shot_medium_cf_fingpt_eval_metrics.txt",
    ("medium", "CF", "counterfactual"): "task/medium/CF/FinGPT/few_shot_cf_medium_cf_fingpt_eval_metrics.txt",
    ("medium", "SE", "zero_shot"):    "task/medium/SE/FinGPT/zero_shot_medium_se_fingpt_eval_metrics.txt",
    ("medium", "SE", "few_shot"):     "task/medium/SE/FinGPT/few_shot_medium_se_fingpt_eval_metrics.txt",
    ("medium", "SE", "counterfactual"): "task/medium/SE/FinGPT/few_shot_cf_medium_se_fingpt_eval_metrics.txt",
    ("medium", "SI", "zero_shot"):    "task/medium/SI/FinGPT/zero_shot_medium_si_fingpt_eval_metrics.txt",
    ("medium", "SI", "few_shot"):     "task/medium/SI/FinGPT/few_shot_medium_si_fingpt_eval_metrics.txt",
    ("medium", "SI", "counterfactual"): "task/medium/SI/FinGPT/few_shot_cf_medium_si_fingpt_eval_metrics.txt",
}

HARD_METRIC_FILES = {
    ("hard", "BS", "zero_shot"):    "task/hard/BS/FinGPT/hard_zero_shot_bs_fingpt_eval_metrics.txt",
    ("hard", "BS", "few_shot"):     "task/hard/BS/FinGPT/hard_few_shot_bs_fingpt_eval_metrics.txt",
    ("hard", "BS", "counterfactual"): "task/hard/BS/FinGPT/hard_cf_few_shot_bs_fingpt_eval_metrics.txt",
    ("hard", "CF", "zero_shot"):    "task/hard/CF/FinGPT/hard_zero_shot_cf_fingpt_eval_metrics.txt",
    ("hard", "CF", "few_shot"):     "task/hard/CF/FinGPT/hard_few_shot_cf_fingpt_eval_metrics.txt",
    ("hard", "CF", "counterfactual"): "task/hard/CF/FinGPT/hard_cf_few_shot_cf_fingpt_eval_metrics.txt",
    ("hard", "SE", "zero_shot"):    "task/hard/SE/FinGPT/hard_zero_shot_se_fingpt_eval_metrics.txt",
    ("hard", "SE", "few_shot"):     "task/hard/SE/FinGPT/hard_few_shot_se_fingpt_eval_metrics.txt",
    ("hard", "SE", "counterfactual"): "task/hard/SE/FinGPT/hard_cf_few_shot_se_fingpt_eval_metrics.txt",
    ("hard", "SI", "zero_shot"):    "task/hard/SI/FinGPT/hard_zero_shot_si_fingpt_eval_metrics.txt",
    ("hard", "SI", "few_shot"):     "task/hard/SI/FinGPT/hard_few_shot_si_fingpt_eval_metrics.txt",
    ("hard", "SI", "counterfactual"): "task/hard/SI/FinGPT/hard_cf_few_shot_si_fingpt_eval_metrics.txt",
}

ALL_METRIC_FILES = {**EASY_METRIC_FILES, **MEDIUM_METRIC_FILES, **HARD_METRIC_FILES}


def parse_float(s: str) -> float:
    """Extract first float from a string like '87.50%' or '87.50'."""
    m = re.search(r"[\d]+\.[\d]+", s)
    return float(m.group()) if m else 0.0


def parse_int_from_line(line: str) -> int:
    m = re.search(r"\d+", line)
    return int(m.group()) if m else 0


def parse_easy_metrics(text: str) -> dict:
    """Parse easy metrics .txt — lines like: Acc:87.50%  Precision:90.00%  Recall:85.00%  F1:87.43%"""
    result = {"accuracy": 0.0, "precision": 0.0, "recall": 0.0, "f1": 0.0, "tokens": 0}
    for line in text.splitlines():
        if re.search(r"Acc[uracy]*:", line, re.I):
            m = re.search(r"Acc[uracy]*:\s*([\d.]+)", line, re.I)
            if m: result["accuracy"] = float(m.group(1))
        if re.search(r"Precision:", line, re.I):
            m = re.search(r"Precision:\s*([\d.]+)", line, re.I)
            if m: result["precision"] = float(m.group(1))
        if re.search(r"Recall:", line, re.I):
            m = re.search(r"Recall:\s*([\d.]+)", line, re.I)
            if m: result["recall"] = float(m.group(1))
        if re.search(r"F1[- ]?Score?:", line, re.I) or re.search(r"\bF1:", line):
            m = re.search(r"F1[- ]?Score?:\s*([\d.]+)", line, re.I) or re.search(r"\bF1:\s*([\d.]+)", line)
            if m: result["f1"] = float(m.group(1))
        if re.search(r"Tokens?:", line, re.I) and ":" in line:
            m = re.search(r"Tokens?:\s*([\d,]+)", line, re.I)
            if m: result["tokens"] = int(m.group(1).replace(",", ""))
    return result


def parse_medium_metrics(text: str) -> dict:
    """Parse medium metrics .txt — lines like: Acc:87.50%  Macro-F1:82.00%"""
    result = {"accuracy": 0.0, "macro_f1": 0.0, "tokens": 0}
    for line in text.splitlines():
        if "Acc" in line and "Macro" not in line:
            m = re.search(r"Acc[uracy]*[:\s]+([\d.]+)", line, re.I)
            if m: result["accuracy"] = float(m.group(1))
        if "Macro" in line and "F1" in line:
            m = re.search(r"Macro-?F1[:\s]+([\d.]+)", line, re.I)
            if m: result["macro_f1"] = float(m.group(1))
        if re.search(r"^Tokens?:", line.strip(), re.I):
            m = re.search(r"Tokens?:\s*([\d,]+)", line, re.I)
            if m: result["tokens"] = int(m.group(1).replace(",", ""))
    return result


def parse_hard_metrics(text: str) -> dict:
    """Parse hard metrics .txt — Step1/Step2 metrics."""
    result = {"step1_accuracy": 0.0, "exact_match": 0.0, "micro_f1": 0.0, "micro_precision": 0.0, "micro_recall": 0.0, "tokens": 0}
    for line in text.splitlines():
        if "Step1" in line and "Acc" in line:
            m = re.search(r"Acc=([\d.]+)%", line)
            if m: result["step1_accuracy"] = float(m.group(1))
        if "ExactMatch" in line:
            m = re.search(r"ExactMatch=([\d.]+)%", line)
            if m: result["exact_match"] = float(m.group(1))
        if "MicroF1" in line:
            m = re.search(r"MicroF1=([\d.]+)%", line)
            if m: result["micro_f1"] = float(m.group(1))
        if "MicroP" in line:
            m = re.search(r"MicroP=([\d.]+)%", line)
            if m: result["micro_precision"] = float(m.group(1))
        if "MicroR" in line:
            m = re.search(r"MicroR=([\d.]+)%", line)
            if m: result["micro_recall"] = float(m.group(1))
        if re.search(r"^Tokens?:", line.strip(), re.I):
            m = re.search(r"Tokens?:\s*([\d,]+)", line, re.I)
            if m: result["tokens"] = int(m.group(1).replace(",", ""))
    return result


def main():
    print("Aggregating FinGPT evaluation results...\n")

    easy_rows = []
    medium_rows = []
    hard_rows = []

    for (difficulty, stmt, strategy), rel_path in ALL_METRIC_FILES.items():
        full_path = REPO_ROOT / rel_path
        if not full_path.exists():
            print(f"  MISSING: {full_path}")
            row_base = {"model": "FinGPT", "difficulty": difficulty, "statement": stmt, "strategy": strategy, "status": "MISSING"}
            if difficulty == "easy":
                easy_rows.append({**row_base, "accuracy": "", "precision": "", "recall": "", "f1": "", "tokens": ""})
            elif difficulty == "medium":
                medium_rows.append({**row_base, "accuracy": "", "macro_f1": "", "tokens": ""})
            else:
                hard_rows.append({**row_base, "step1_accuracy": "", "exact_match": "", "micro_precision": "", "micro_recall": "", "micro_f1": "", "tokens": ""})
            continue

        text = full_path.read_text(encoding="utf-8")
        print(f"  OK: {rel_path}")

        if difficulty == "easy":
            m = parse_easy_metrics(text)
            easy_rows.append({
                "model": "FinGPT", "difficulty": difficulty, "statement": stmt, "strategy": strategy,
                "accuracy": f"{m['accuracy']:.2f}", "precision": f"{m['precision']:.2f}",
                "recall": f"{m['recall']:.2f}", "f1": f"{m['f1']:.2f}", "tokens": m["tokens"], "status": "OK"
            })
        elif difficulty == "medium":
            m = parse_medium_metrics(text)
            medium_rows.append({
                "model": "FinGPT", "difficulty": difficulty, "statement": stmt, "strategy": strategy,
                "accuracy": f"{m['accuracy']:.2f}", "macro_f1": f"{m['macro_f1']:.2f}",
                "tokens": m["tokens"], "status": "OK"
            })
        else:
            m = parse_hard_metrics(text)
            hard_rows.append({
                "model": "FinGPT", "difficulty": difficulty, "statement": stmt, "strategy": strategy,
                "step1_accuracy": f"{m['step1_accuracy']:.2f}", "exact_match": f"{m['exact_match']:.2f}",
                "micro_precision": f"{m['micro_precision']:.2f}", "micro_recall": f"{m['micro_recall']:.2f}",
                "micro_f1": f"{m['micro_f1']:.2f}", "tokens": m["tokens"], "status": "OK"
            })

    # Write CSVs
    easy_csv = REPO_ROOT / "fingpt_easy_results_summary.csv"
    medium_csv = REPO_ROOT / "fingpt_medium_results_summary.csv"
    hard_csv = REPO_ROOT / "fingpt_hard_results_summary.csv"

    easy_fields = ["model", "difficulty", "statement", "strategy", "accuracy", "precision", "recall", "f1", "tokens", "status"]
    medium_fields = ["model", "difficulty", "statement", "strategy", "accuracy", "macro_f1", "tokens", "status"]
    hard_fields = ["model", "difficulty", "statement", "strategy", "step1_accuracy", "exact_match", "micro_precision", "micro_recall", "micro_f1", "tokens", "status"]

    def write_csv(path, rows, fields):
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=fields)
            w.writeheader(); w.writerows(rows)

    write_csv(easy_csv, easy_rows, easy_fields)
    write_csv(medium_csv, medium_rows, medium_fields)
    write_csv(hard_csv, hard_rows, hard_fields)

    # Print consolidated summary table
    print("\n" + "="*120)
    print("FINGPT EVALUATION SUMMARY — FinGPT/fingpt-mt_llama2-7b_lora")
    print("="*120)

    for difficulty, rows, header_keys in [
        ("EASY", easy_rows, ["accuracy", "f1"]),
        ("MEDIUM", medium_rows, ["accuracy", "macro_f1"]),
        ("HARD", hard_rows, ["step1_accuracy", "exact_match", "micro_f1"]),
    ]:
        print(f"\n{'─'*120}")
        print(f"  {difficulty} TASK")
        print(f"{'─'*120}")
        col_header = f"  {'Statement':<6} {'Strategy':<16} " + "  ".join(f"{k.upper():<14}" for k in header_keys) + "  Tokens"
        print(col_header)
        print(f"  {'─'*110}")
        for row in rows:
            vals = "  ".join(f"{row.get(k,'N/A'):<14}" for k in header_keys)
            tok = str(row.get("tokens",""))
            status = "" if row.get("status")=="OK" else " [MISSING]"
            print(f"  {row['statement']:<6} {row['strategy']:<16} {vals}  {tok}{status}")

    print(f"\n{'='*120}")
    print(f"Output CSVs:")
    print(f"  Easy:   {easy_csv}")
    print(f"  Medium: {medium_csv}")
    print(f"  Hard:   {hard_csv}")

    # Try to write Excel if openpyxl available
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()

        def fill_sheet(ws, rows, fields, title):
            ws.title = title
            ws.append(fields)
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="4472C4")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center")
            for row in rows:
                ws.append([row.get(f, "") for f in fields])
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col) + 2
                ws.column_dimensions[col[0].column_letter].width = min(max_len, 25)

        fill_sheet(wb.active, easy_rows, easy_fields, "Easy")
        fill_sheet(wb.create_sheet(), medium_rows, medium_fields, "Medium")
        fill_sheet(wb.create_sheet(), hard_rows, hard_fields, "Hard")

        xlsx_path = REPO_ROOT / "fingpt_results_summary.xlsx"
        wb.save(xlsx_path)
        print(f"  Excel:  {xlsx_path}")
    except ImportError:
        print("  (Install openpyxl for Excel output: pip install openpyxl)")

    print("="*120)


if __name__ == "__main__":
    main()
